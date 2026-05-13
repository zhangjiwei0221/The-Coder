from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT_XLSX = ROOT / "怪物配置表.xlsx"
OUT_MD = ROOT / "怪物配置表.md"


ENEMY_ROWS = [
    ["bug", "BUG", "第一层", "普通/教学", 28, "攻击[6] / 防御[5]", "行动循环: 攻击 -> 防御 -> 攻击+防御", "第一层基础攻防教学，验证敌方代码、多行意图和护盾窗口", "img/enemies/bug.png", "数值已实装；新素材待接入"],
    ["error", "Error", "第一层", "普通", 30, "攻击[7-10]", "每回合只攻击；每次受到攻击有 10% 概率下回合眩晕", "高一点的随机攻击压力，鼓励玩家主动攻击并理解打断机会", "img/enemies/error.png", "素材已生成；机制待接入"],
    ["loop", "回环", "第一层", "普通", 32, "①循环(3){攻击[3]} / ②循环(2){攻击[3] + 防御[3]}", "通过固定多段行动引入循环机制；第二段循环同时攻击与防御", "让玩家第一次理解循环不是抽象语法，而是重复执行的敌方行动", "img/enemies/loop.png", "素材已生成；机制待接入"],
    ["branch", "岔路", "第一层", "普通", 34, "①攻击[3]；如果(你没有护盾) 攻击[5]；如果(自身HP<30%) 攻击[3] / ②如果(你有护盾) 防御[5]；否则 下次攻击参数+1", "根据玩家护盾与自身血量走不同分支；无护盾时追击，低血量时补刀，有护盾时转防御，否则蓄积参数", "教玩家用防御影响敌方分支，建立 if/else 的战斗直觉", "img/enemies/branch.png", "素材已生成；机制待接入"],
    ["nullptr", "空值", "第一层", "精英候选", 38, "攻击[7] / // 空值潜伏 / 攻击[11]", "潜伏一回合后打出较高伤害", "后段或精英节点的节奏预判测试，暂不放入前段普通池", "", "旧配置保留；待重命名/重做素材"],
    ["chaser", "催单", "第一层", "精英候选", 40, "攻击[4+t*2]", "伤害随回合递增", "拖延惩罚，逼玩家尽快结束战斗", "", "旧配置保留；待重做素材"],
    ["todo", "待办", "第一层", "精英候选", 40, "攻击[8] / 防御[7]", "攻击与防御交替", "教玩家识别敌方防御回合和输出窗口", "", "旧配置保留；待重做素材"],
    ["infloop", "InfiniteLoop", "第二层", "普通", 35, "攻击[3+t*2]", "伤害随回合线性成长", "第二层成长压力怪", "", "待重命名"],
    ["memleak", "MemoryLeak", "第二层", "普通", 40, "吸血[4-6]", "造成伤害后回复一半生命", "拖延惩罚，要求稳定输出", "", "待重命名"],
    ["recursion", "Recursion", "第二/三层", "精英", 45, "攻击[5,10,20,40]", "伤害指数增长，上限 40", "强爆发预警怪", "", "待重命名"],
    ["stackoverflow", "StackOverflow", "第二/三层", "精英", 55, "循环(3){攻击[4]}", "多段攻击，总伤害 12", "检验护盾、反弹和多段防御能力", "", "待重命名"],
    ["racecond", "RaceCondition", "第二/三层", "普通", 50, "攻击[3-20] // 随机", "伤害大幅随机，但会提前显示本回合数值", "高波动风险怪", "", "待重命名"],
    ["deadlock", "Deadlock", "第二/三层", "精英", 60, "防御[15] / 攻击[20]", "防御与重击交替", "窗口型战斗，鼓励破盾和爆发", "", "待重命名"],
    ["gc", "GarbageCollector", "第二/三层", "普通", 35, "垃圾回收() / 攻击[6-10]", "偶数回合移除 1 张手牌", "手牌资源干扰", "", "待重命名"],
    ["syntaxerr", "红字", "第一层", "Boss", 125, "防御[10] / 攻击[10-15] / 语法错误!", "周期性护盾、攻击、随机弃 1 张手牌", "第一层综合检验 Boss", "", "已实装；待重做素材"],
    ["firewall", "Firewall", "第二层", "Boss", 120, "防御[20] / 防御[15]+攻击[10] / 攻击[25]", "高护盾与高爆发交替", "第二层护盾压力 Boss", "", "待重命名"],
    ["root", "Root", "第三层", "Boss", 200, "循环(3){防御[10]} / 治疗[20] / 攻击[20-30]", "护盾、治疗、高伤害循环", "最终综合 Boss", "", "待重命名"],
]

FIRST_LAYER_ROWS = [
    ["BUG", "第一层前段普通池", "低", "HP 28；攻击/防御/攻击+防御循环，作为基础攻防教学"],
    ["Error", "第一层前段普通池", "中低", "HP 30；攻击 7-10；受击 10% 概率下回合眩晕，压力稍高但有打断机会"],
    ["回环", "第一层第 3-4 列后加入", "中低", "HP 32；①循环 3 次攻击 3；②循环 2 次攻击 3 + 防御 3，引入循环机制"],
    ["岔路", "第一层第 3-4 列后加入", "中", "HP 34；①先攻击 3，再按玩家护盾/自身低血量判断追加攻击；②根据玩家是否有护盾选择防御或下次攻击参数+1"],
    ["空值", "第一层精英候选", "中", "HP 38；潜伏后攻击，适合教预判；暂不进入前段普通池"],
    ["催单", "第一层精英候选", "中", "HP 40；回合越久越危险"],
    ["待办", "第一层精英候选", "中", "HP 40；防御回合给玩家观察窗口"],
    ["红字", "第一层 Boss", "中高", "HP 125；目标 6-8 回合战斗；后续需统一新美术风格"],
]

BASELINE_ROWS = [
    ["普通怪 HP", "28-34", "第一层普通战斗控制在约 3-4 回合，优先教学而非消耗"],
    ["普通怪常规攻击", "5-7", "玩家需要考虑防御，但不应被早期连续压血"],
    ["普通怪压力攻击", "7-10", "放在 Error 或带条件的岔路中，必须能通过主动攻击/防御缓解"],
    ["循环/判断引入", "第 3-4 列后", "先让玩家理解基础攻防，再引入 Loop 和 If 的敌方机制"],
    ["第一层精英 HP", "38-40", "普通战不再抽精英怪，精英节点承担节奏测试"],
    ["第一层 Boss HP", "125", "目标 6-8 回合"],
    ["第一层 Boss 攻击", "10-15", "覆盖防御检查和血量压力"],
    ["普通战斗目标时长", "3-4 回合", "让机制有时间出现"],
    ["Boss 战目标时长", "6-8 回合", "第一层综合检验"],
]

FIELD_ROWS = [
    ["字段", "说明"],
    ["ID", "代码中的怪物键名，后续用于表格和程序对齐"],
    ["怪物名", "玩家在游戏中看到的名称"],
    ["层级", "怪物主要出现的层级"],
    ["类型", "普通、精英、Boss 或教学"],
    ["HP", "当前生命值配置"],
    ["代码库", "右侧代码库展示内容，代表敌人的行动模板"],
    ["特殊机制", "该怪物区别于普通攻击怪的关键规则"],
    ["设计定位", "它负责教玩家什么，或制造哪类压力"],
    ["素材路径", "透明 PNG 或后续美术资源路径"],
    ["状态", "已实装、待重命名、待调整等"],
]


def style_sheet(ws, freeze="A2"):
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(name="微软雅黑", bold=True, color="000000")
    body_font = Font(name="微软雅黑", size=10, color="000000")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            cell.font = body_font
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 15, "B": 14, "C": 12, "D": 13, "E": 9,
        "F": 36, "G": 38, "H": 38, "I": 24, "J": 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_rows(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.auto_filter.ref = ws.dimensions
    style_sheet(ws)


def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "怪物配置总表"
    add_rows(ws, ["ID", "怪物名", "层级", "类型", "HP", "代码库", "特殊机制", "设计定位", "素材路径", "状态"], ENEMY_ROWS)

    ws = wb.create_sheet("第一层配置")
    add_rows(ws, ["怪物", "推荐出现场景", "机制强度", "备注"], FIRST_LAYER_ROWS)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 56

    ws = wb.create_sheet("第一层数值基准")
    add_rows(ws, ["项目", "当前目标", "说明"], BASELINE_ROWS)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 52

    ws = wb.create_sheet("字段说明")
    add_rows(ws, ["字段", "说明"], FIELD_ROWS[1:])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 70

    wb.save(OUT_XLSX)


def build_markdown():
    headers = ["ID", "怪物名", "层级", "类型", "HP", "代码库", "特殊机制", "设计定位", "素材路径", "状态"]
    lines = ["# 怪物配置表", "", "| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
    for row in ENEMY_ROWS:
        lines.append("| " + " | ".join(str(v).replace("\n", "<br>") for v in row) + " |")
    lines.extend(["", "## 第一层数值基准", "", "| 项目 | 当前目标 | 说明 |", "| --- | --- | --- |"])
    for row in BASELINE_ROWS:
        lines.append("| " + " | ".join(str(v) for v in row) + " |")
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


if __name__ == "__main__":
    build_workbook()
    build_markdown()
    print(OUT_XLSX)
    print(OUT_MD)
